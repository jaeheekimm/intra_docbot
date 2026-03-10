import os
import logging
from typing import Dict, List, Any, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from langchain_core.documents import Document

from src.utils.files import safe_filename
from src.utils.hashing import sha1_short

logger = logging.getLogger(__name__)

SHEET_MODE_THRESHOLD = 50  # 비어있지 않은 행이 이것보다 적으면 시트 단위로
SHEET_MODE_KEYWORDS = [
    "신청서",
    "양식",
    "서식",
    "신청",
    "지원서",
]  # 파일명에 이 키워드 있으면 무조건 시트 단위


def _guess_image_ext(img) -> str:
    """openpyxl Image 객체에서 확장자를 추정. 실패하면 'bin' 반환.

    openpyxl 버전마다 속성이 달라서 format → _format → path 순으로 시도.
    """
    fmt = getattr(img, "format", None) or getattr(img, "_format", None)
    if isinstance(fmt, str) and fmt.strip():
        return fmt.strip().lower().lstrip(".")

    path = getattr(img, "path", None)
    if isinstance(path, str) and "." in os.path.basename(path):
        ext = os.path.splitext(path)[1].lower().lstrip(".")
        if ext:
            return ext

    return "bin"


def extract_xlsx_images(xlsx_path: str, out_dir: str) -> Dict[str, List[str]]:
    """XLSX에서 이미지 추출 후 out_dir에 저장.

    반환: {시트이름: [이미지경로, ...]}
    """
    mapping: Dict[str, List[str]] = {}
    wb = load_workbook(xlsx_path, data_only=True)
    base = safe_filename(os.path.basename(xlsx_path))

    for ws in wb.worksheets:
        imgs = getattr(ws, "_images", []) or []
        if not imgs:
            continue

        for i, img in enumerate(imgs, start=1):
            try:
                blob = img._data()
            except Exception as e:
                logger.warning(
                    "Failed to extract image: file=%s sheet=%s idx=%s err=%s",
                    xlsx_path,
                    ws.title,
                    i,
                    repr(e),
                )
                continue

            if not blob:
                logger.warning(
                    "Empty image blob: file=%s sheet=%s idx=%s",
                    xlsx_path,
                    ws.title,
                    i,
                )
                continue

            hid = sha1_short(blob)
            ext = _guess_image_ext(img)
            fname = f"{base}__sheet{safe_filename(ws.title)}__img{i:03d}__{hid}.{ext}"
            save_path = os.path.join(out_dir, fname)

            if not os.path.exists(save_path):
                with open(save_path, "wb") as f:
                    f.write(blob)

            mapping.setdefault(ws.title, []).append(save_path)

    wb.close()
    return mapping


def _pick_header_row(ws, max_scan_rows: int = 20) -> Tuple[int, List[str]]:
    """상단 max_scan_rows 행 중 헤더로 쓸 첫 번째 비어있지 않은 행을 찾음.

    - 빈 셀은 A, B, C... 컬럼 레터로 대체
    반환: (헤더 행 인덱스(1-based), 헤더 리스트)
    """
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > max_scan_rows:
            break

        vals = ["" if v is None else str(v).strip() for v in row]
        if all(v == "" for v in vals):
            continue

        headers: List[str] = []
        for c_idx, v in enumerate(vals, start=1):
            h = v if v != "" else get_column_letter(c_idx)
            headers.append(h)
        return r_idx, headers

    return 0, []


def _row_to_text_with_headers(headers: List[str], values: Tuple[Any, ...]) -> str:
    """행 값을 헤더 기반으로 "헤더: 값 | 헤더: 값" 형태 문자열로 변환"""
    parts: List[str] = []
    for i, v in enumerate(values):
        if i >= len(headers):
            break
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        parts.append(f"{headers[i]}: {s}")
    return " | ".join(parts).strip()


def _expand_merged_cells(ws):
    """병합 셀을 해제하고 병합 전 값을 모든 셀에 복사.

    병합 상태에서 바로 unmerge하면 값이 사라지므로,
    먼저 값을 저장해두고 해제 후 다시 채워넣는 순서가 중요.
    """
    merge_data = []
    for merge_range in list(ws.merged_cells.ranges):
        top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
        merge_data.append(
            (
                merge_range.min_row,
                merge_range.max_row,
                merge_range.min_col,
                merge_range.max_col,
                top_left_value,
            )
        )

    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))

    for min_row, max_row, min_col, max_col, value in merge_data:
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).value = value

    return ws


def load_xlsx_docs_rows(
    xlsx_path: str,
    xlsx_img_map: Dict[str, List[str]],
    *,
    attach_sheet_images: bool = False,
    max_rows_per_sheet: Optional[int] = None,
    header_scan_rows: int = 20,
) -> List[Document]:
    """XLSX를 Document 리스트로 변환.

    시트 단위 모드: 파일명에 신청서/양식 등 키워드 있거나 비어있지 않은 행이 50개 미만인 경우.
    행 단위 모드: 데이터가 많은 일반 표.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    file_name = os.path.basename(xlsx_path)
    docs: List[Document] = []

    force_sheet_mode = any(kw in file_name for kw in SHEET_MODE_KEYWORDS)

    for ws in wb.worksheets:
        ws = _expand_merged_cells(ws)

        sheet_images = xlsx_img_map.get(ws.title, [])
        image_count = len(sheet_images)

        non_empty_rows = [
            row
            for row in ws.iter_rows(values_only=True)
            if any(v is not None and str(v).strip() for v in row)
        ]
        row_count_actual = len(non_empty_rows)

        if force_sheet_mode or row_count_actual < SHEET_MODE_THRESHOLD:
            # ── 시트 단위 모드 (복잡한 양식/병합 표에 적합) ────────────────

            lines = []
            for row in non_empty_rows:
                vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if vals:
                    # 병합 셀 확장으로 생긴 행 내 중복 값 제거 (순서 유지)
                    seen_set = set()
                    unique_vals = []
                    for v in vals:
                        if v not in seen_set:
                            unique_vals.append(v)
                            seen_set.add(v)
                    lines.append(" | ".join(unique_vals))

            seen_lines = set()
            unique_lines = []
            for line in lines:
                if line not in seen_lines:
                    unique_lines.append(line)
                    seen_lines.add(line)

            text = "\n".join(unique_lines).strip()
            if not text:
                continue

            docs.append(
                Document(
                    page_content=text,
                    metadata={
                        "source": xlsx_path,
                        "file_name": file_name,
                        "doc_type": "xlsx",
                        "page": None,
                        "slide": None,
                        "sheet": ws.title,
                        "row": None,
                        "image_count": image_count,
                        "image_paths": (sheet_images if attach_sheet_images else []),
                    },
                )
            )

        else:
            # ── 행 단위 모드 (데이터가 많은 일반 표에 적합) ──────────────────
            header_row_idx, headers = _pick_header_row(
                ws, max_scan_rows=header_scan_rows
            )

            row_count = 0
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if header_row_idx and r_idx == header_row_idx:
                    continue

                if headers:
                    text = _row_to_text_with_headers(headers, row)
                else:
                    vals = ["" if v is None else str(v).strip() for v in row]
                    if all(v == "" for v in vals):
                        continue
                    text = " | ".join([v for v in vals if v != ""]).strip()

                if not text:
                    continue

                docs.append(
                    Document(
                        page_content=text,
                        metadata={
                            "source": xlsx_path,
                            "file_name": file_name,
                            "doc_type": "xlsx",
                            "page": None,
                            "slide": None,
                            "sheet": ws.title,
                            "row": r_idx,
                            "image_count": image_count,
                            "image_paths": (
                                sheet_images if attach_sheet_images else []
                            ),
                        },
                    )
                )

                row_count += 1
                if max_rows_per_sheet and row_count >= max_rows_per_sheet:
                    break

            # 텍스트 없고 이미지만 있으면 빈 content로 Doc 생성 (이미지 참조용)
            if row_count == 0 and sheet_images:
                docs.append(
                    Document(
                        page_content="",
                        metadata={
                            "source": xlsx_path,
                            "file_name": file_name,
                            "doc_type": "xlsx",
                            "page": None,
                            "slide": None,
                            "sheet": ws.title,
                            "row": None,
                            "image_count": image_count,
                            "image_paths": sheet_images,
                        },
                    )
                )

    wb.close()
    return docs


def xlsx_image_manifest(
    xlsx_path: str, xlsx_img_map: Dict[str, List[str]]
) -> List[Dict[str, Any]]:
    """시트별 이미지 경로 정보를 image_manifest 포맷의 딕셔너리 리스트로 반환"""
    items: List[Dict[str, Any]] = []
    for sheet, paths in xlsx_img_map.items():
        for ip in paths:
            items.append(
                {
                    "source_file": xlsx_path,
                    "source_type": "xlsx",
                    "page": None,
                    "slide": None,
                    "sheet": sheet,
                    "row": None,
                    "image_path": ip,
                }
            )
    return items
